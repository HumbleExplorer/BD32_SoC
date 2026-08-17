"""
BD32 CoreMark 自动化测试脚本
流程：复位 → UART下载程序 → 等待输出 → 解析结果

用法：
    python tools/auto_coremark.py                          # 跑全部23个配置
    python tools/auto_coremark.py --compiler gcc --opt o2  # 只跑 GCC O2
    python tools/auto_coremark.py --compiler clang         # 只跑 LLVM/Clang 全部配置
    python tools/auto_coremark.py --libc picolibc          # 只跑 picolibc 配置
    python tools/auto_coremark.py --lld                    # picolibc LLVM 5 档改用 LTO+ICF 变体
    python tools/auto_coremark.py --port COM8 --baud 115200

结果自动写入 BD32_CoreMark_Compiler_Comparison.xlsx（--xlsx 可改路径）。
ITCM/DTCM 词数从 .uartbin 帧头解析（START + ITCM_CNT + ITCM + DTCM_CNT + DTCM）。
"""
import subprocess
import serial
import serial.tools.list_ports
import time
import struct
import argparse
import os
import sys
import re

# ============================================================
# 配置
# ============================================================
# 复位由 fpga_reset.py 处理（OpenOCD/WinUSB，无需 ftd2xx）
BOOT_DELAY = 1.0        # 复位释放后等BootROM启动(秒)
SEND_CHUNK = 256        # 串口发送分块大小
SEND_DELAY = 0.01       # 每块间隔(秒)
RESULT_TIMEOUT = 300    # 等待结果超时(秒), CoreMark 500轮约需3-4分钟

# 23个测试配置: (名称, uartbin文件名)
# 18 个编译器/优化/C库配置（GCC/LLVM x Os/O1/O2/O3/Oz x newlib-nano/picolibc）
# + 5 个 picolibc LLVM 的 LLD（LTO+ICF）变体
CONFIGS = [
    # ---- newlib-nano ----
    ("GCC Os",   "coremark_os.uartbin"),
    ("GCC O1",   "coremark_o1.uartbin"),
    ("GCC O2",   "coremark_o2.uartbin"),
    ("GCC O3",   "coremark_o3.uartbin"),
    ("LLVM Oz",  "coremark_clang_oz.uartbin"),
    ("LLVM Os",  "coremark_clang_os.uartbin"),
    ("LLVM O1",  "coremark_clang_o1.uartbin"),
    ("LLVM O2",  "coremark_clang_o2.uartbin"),
    ("LLVM O3",  "coremark_clang_o3.uartbin"),
    # ---- picolibc（整数档）----
    ("Pico GCC Os",   "coremark_picolibc_os.uartbin"),
    ("Pico GCC O1",   "coremark_picolibc_o1.uartbin"),
    ("Pico GCC O2",   "coremark_picolibc_o2.uartbin"),
    ("Pico GCC O3",   "coremark_picolibc_o3.uartbin"),
    ("Pico LLVM Oz",  "coremark_picolibc_clang_oz.uartbin"),
    ("Pico LLVM Os",  "coremark_picolibc_clang_os.uartbin"),
    ("Pico LLVM O1",  "coremark_picolibc_clang_o1.uartbin"),
    ("Pico LLVM O2",  "coremark_picolibc_clang_o2.uartbin"),
    ("Pico LLVM O3",  "coremark_picolibc_clang_o3.uartbin"),
    # ---- picolibc LLVM + LLD（LTO+ICF）变体 ----
    ("Pico LLVM Oz (lld)",  "coremark_picolibc_clang_lld_oz.uartbin"),
    ("Pico LLVM Os (lld)",  "coremark_picolibc_clang_lld_os.uartbin"),
    ("Pico LLVM O1 (lld)",  "coremark_picolibc_clang_lld_o1.uartbin"),
    ("Pico LLVM O2 (lld)",  "coremark_picolibc_clang_lld_o2.uartbin"),
    ("Pico LLVM O3 (lld)",  "coremark_picolibc_clang_lld_o3.uartbin"),
]


def parse_uartbin_size(path):
    """从 .uartbin 帧头解析 ITCM/DTCM 词数。
    帧格式：START_FRAME(4) + ITCM_CNT(4) + ITCM words + DTCM_CNT(4) + DTCM words"""
    try:
        with open(path, "rb") as f:
            data = f.read()
        if len(data) < 12:
            return 0, 0
        itcm_cnt = struct.unpack("<I", data[4:8])[0]
        off = 8 + itcm_cnt * 4
        if off + 4 > len(data):
            return itcm_cnt, 0
        dtcm_cnt = struct.unpack("<I", data[off:off + 4])[0]
        return itcm_cnt, dtcm_cnt
    except OSError:
        return 0, 0

# ============================================================
# 串口自动检测
# ============================================================
def find_port():
    """自动检测 CH340 USB-UART 串口"""
    candidates = []
    for p in serial.tools.list_ports.comports():
        desc = (p.description or "").upper()
        vid_pid = f"{p.vid:04X}:{p.pid:04X}" if p.vid else ""
        if "CH340" in desc or vid_pid == "1A86:7523":
            candidates.append(p.device)
    if len(candidates) == 1:
        return candidates[0]
    elif len(candidates) > 1:
        print(f"[WARN] Multiple CH340 found: {candidates}, using {candidates[0]}")
        return candidates[0]
    return None

# ============================================================
# 复位控制
# ============================================================
def reset_fpga():
    """复位 FPGA：调用 fpga_reset.py（默认 OpenOCD/WinUSB，无需切换驱动）"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    r = subprocess.run([sys.executable, os.path.join(script_dir, "fpga_reset.py")],
                       capture_output=True, text=True, errors="replace")
    if r.returncode != 0:
        print("[ERROR] FPGA reset failed:\n%s" % ((r.stderr or r.stdout)[-400:]))
        return False
    return True

# ============================================================
# UART 下载
# ============================================================
def send_program(ser, filepath):
    data = open(filepath, 'rb').read()
    print(f"    Sending {len(data)} bytes ({len(data)//SEND_CHUNK+1} chunks)...")
    for i in range(0, len(data), SEND_CHUNK):
        chunk = data[i:i+SEND_CHUNK]
        ser.write(chunk)
        time.sleep(SEND_DELAY)
    ser.flush()
    print(f"    Send complete.")

# ============================================================
# 等待并解析结果
# ============================================================
def wait_for_result(ser, timeout=RESULT_TIMEOUT):
    """等待 CoreMark 输出完成，返回全部输出文本"""
    print(f"    Waiting for CoreMark output (timeout={timeout}s)...")
    output = b""
    start = time.time()
    done = False

    while time.time() - start < timeout:
        if ser.in_waiting > 0:
            chunk = ser.read(ser.in_waiting)
            output += chunk
            text = output.decode('utf-8', errors='replace')
            # 检测完成标志
            if "Correct operation validated" in text or "CoreMark/MHz" in text:
                # 多等2秒确保输出完整
                time.sleep(2)
                if ser.in_waiting > 0:
                    output += ser.read(ser.in_waiting)
                done = True
                break
        else:
            time.sleep(0.1)

    text = output.decode('utf-8', errors='replace')
    if not done:
        print(f"    [WARN] Timeout! Got {len(output)} bytes so far.")
    return text, done

def parse_coremark(text):
    """从输出文本中解析 CoreMark 结果"""
    result = {}

    m = re.search(r'CoreMark/MHz\s*[:\s]+([\d.]+)', text)
    if m: result['coremark_mhz'] = float(m.group(1))

    m = re.search(r'Iterations/Sec\s*[:\s]+([\d.]+)', text)
    if m: result['iter_per_sec'] = float(m.group(1))

    m = re.search(r'Total ticks\s*[:\s]+(\d+)', text)
    if m: result['total_ticks'] = int(m.group(1))

    m = re.search(r'Predict rate\s*:\s*([0-9.]+)\s*%', text)
    if m: result['branch_hit'] = float(m.group(1)) / 100.0

    return result


def write_xlsx(configs, results, xlsx_path):
    """把测试结果写入 Excel：主表（23 配置）+ 归一化对比 + Charts 数据。
      体积比/性能比等以各 C 库的 GCC O2 为基准，用公式随主表联动。"""
    import openpyxl
    from openpyxl.styles import Font

    # 若传入的是目录（如 --xlsx ../../），视为输出目录，拼默认文件名
    if os.path.isdir(xlsx_path):
        xlsx_path = os.path.join(xlsx_path, "BD32_CoreMark_Compiler_Comparison.xlsx")

    # 收集有结果的行
    data = []
    for name, fname in configs:
        r = results.get(name)
        if not r:
            continue
        is_lld = " (lld)" in name
        parts = name.replace(" (lld)", "").split()
        libc = "picolibc" if parts[0] == "Pico" else "newlib-nano"
        compiler = parts[1] if parts[0] == "Pico" else parts[0]
        opt_raw = parts[-1]
        opt = "-" + opt_raw[:1] + opt_raw[1:].lower()   # Os -> -Os, Oz -> -Oz
        data.append((name, compiler, libc, opt, r, is_lld))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BD32编译器优化对比"
    bold = Font(bold=True)

    # ---- 标题 / 参数 ----
    ws["A1"] = "BD32 RV32IM CoreMark 编译器优化对比表"
    ws["A2"] = ("CPU: BD32 5-stage pipeline | Freq: 75MHz | CoreMark Size: 2000 (2K) | "
                "Iterations: 500 | march=rv32im_zicsr | mabi=ilp32 | 链接: xPack GCC/GNU ld")
    ws["A3"] = ("picolibc = picolibc 1.8.12 整数档 (format-default=i)；"
                "数据由 auto_coremark.py 自动采集（体积来自 .uartbin 帧头）")

    headers = ["序号", "编译器", "C库", "优化选项", "ITCM词数(words)", "DTCM词数(words)",
               "总代码体积(bytes)", "CoreMark/MHz", "Iterations/Sec", "Total ticks",
               "分支预测成功率(%)", "备注"]
    for j, h in enumerate(headers, 1):
        ws.cell(4, j, h).font = bold

    # ---- 主表（行 5 起）----
    base_rows = {}   # libc -> 主表行号（GCC O2 基准）
    for i, (name, compiler, libc, opt, r, is_lld) in enumerate(data):
        row = 5 + i
        ws.cell(row, 1, i + 1)
        ws.cell(row, 2, compiler)
        ws.cell(row, 3, libc)
        ws.cell(row, 4, opt + "  -march=rv32im_zicsr -mabi=ilp32")
        ws.cell(row, 5, r.get("itcm_words", ""))
        ws.cell(row, 6, r.get("dtcm_words", ""))
        ws.cell(row, 7, f"=(E{row}+F{row})*4")
        ws.cell(row, 8, r.get("coremark_mhz", ""))
        ws.cell(row, 9, r.get("iter_per_sec", ""))
        ws.cell(row, 10, r.get("total_ticks", ""))
        br = r.get("branch_hit")
        ws.cell(row, 11, round(br * 100, 2) if br else "")
        if opt == "-O2" and compiler == "GCC" and libc == "newlib-nano":
            ws.cell(row, 12, "基准(newlib)")
            base_rows["newlib-nano"] = row
        elif opt == "-O2" and compiler == "GCC" and libc == "picolibc":
            ws.cell(row, 12, "基准(picolibc)")
            base_rows["picolibc"] = row
        elif libc == "picolibc":
            ws.cell(row, 12, "picolibc 1.8.12 整数档" + (" + LLD/LTO/ICF" if is_lld else ""))

    # ---- 归一化对比（以各 C 库 GCC O2 = 1.00）----
    norm_start = 5 + len(data) + 2
    ws.cell(norm_start, 1, "归一化对比（以各 C 库 GCC -O2 为基准 = 1.00）").font = bold
    norm_hdr = norm_start + 1
    for j, h in enumerate(["序号", "编译器", "C库", "优化等级", "体积比(vs O2)",
                           "性能比(vs O2)", "性价比(性能/体积)"], 1):
        ws.cell(norm_hdr, j, h).font = bold
    for i, (name, compiler, libc, opt, r, is_lld) in enumerate(data):
        src_row = 5 + i                 # 主表行
        dst_row = norm_hdr + 1 + i
        base = base_rows.get(libc)
        ws.cell(dst_row, 1, i + 1)
        ws.cell(dst_row, 2, compiler)
        ws.cell(dst_row, 3, libc)
        ws.cell(dst_row, 4, opt)
        ws.cell(dst_row, 5, f'=IF(G${base}=0,"",G{src_row}/G${base})' if base else "")
        ws.cell(dst_row, 6, f'=IF(H${base}=0,"",H{src_row}/H${base})' if base else "")
        # 性价比 = 性能比 / 体积比（E 体积比、F 性能比），越大越好
        ws.cell(dst_row, 7, '=IF(OR(E{d}="",E{d}=0,F{d}="",F{d}=0),"",F{d}/E{d})'.format(d=dst_row))

    # ---- 说明 ----
    note_start = norm_hdr + 1 + len(data) + 1
    notes = [
        "说明：",
        "1. ITCM/DTCM 词数从 .uartbin 帧头解析（START_FRAME + ITCM_CNT + ITCM + DTCM_CNT + DTCM）。",
        "2. CoreMark/MHz、Iterations/Sec、Total ticks、分支预测成功率由固件 UART 输出解析。",
        "3. 归一化对比自动计算（体积/性能比引用主表，只写主表数据即可联动）。",
        "4. GCC 工具链：riscv-none-elf-gcc（xPack, GCC 15.2.0）；LLVM：clang 22.1.8 前端，经 xPack GCC 驱动链接。",
        "5. picolibc = picolibc 1.8.12 整数 printf 档（format-default=i）；浮点档体积明显更大，非本表范围。",
        "6. LLVM LTO+ICF（--lld）可进一步缩小体积，数据见 doc/sdk.md「LLVM/Clang 集成与代码体积优化」。",
    ]
    for k, note in enumerate(notes):
        ws.cell(note_start + k, 1, note)

    # ---- Charts 数据 ----
    wc = wb.create_sheet("Charts")
    chart_keys = ["GCC", "LLVM", "Pico GCC", "Pico LLVM", "Pico LLVM+LLD"]
    wc.append(["Opt"] + [f"{k} CM/MHz" for k in chart_keys] + [""] +
              [f"{k} 体积" for k in chart_keys] + [""] +
              [f"{k} 分支%" for k in chart_keys])
    for opt in ["Oz", "Os", "O1", "O2", "O3"]:
        o = "-" + opt[:1] + opt[1:].lower()
        row = {k: None for k in chart_keys}
        for name, compiler, libc, optx, r, is_lld in data:
            if optx != o:
                continue
            key = ("Pico " if libc == "picolibc" else "") + compiler + ("+LLD" if is_lld else "")
            row[key] = r
        cm = [row[k]["coremark_mhz"] if row[k] else None for k in chart_keys]
        sz = [((row[k].get("itcm_words", 0) + row[k].get("dtcm_words", 0)) * 4)
              if row[k] else None for k in chart_keys]
        br = [round(row[k]["branch_hit"] * 100, 2) if row[k] and row[k].get("branch_hit")
              else None for k in chart_keys]
        wc.append([opt] + cm + [None] + sz + [None] + br)

    try:
        wb.save(xlsx_path)
        print(f"\n[XLSX] results written -> {xlsx_path}")
    except PermissionError:
        # 目标文件被 WPS/Excel 占用时，写带时间戳的备用文件，避免测试结果丢失
        import time as _time
        base = os.path.basename(xlsx_path) or "BD32_CoreMark_Compiler_Comparison.xlsx"
        stem = os.path.splitext(base)[0] or "BD32_CoreMark_Compiler_Comparison"
        alt = os.path.join(os.path.dirname(xlsx_path),
                           f"{stem}_{_time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(alt)
        # 用 Office/WPS 锁文件（~$<文件名>）判断是否真被 WPS/Excel 打开，
        # 避免"开着 WPS 但没打开该文件"时误报
        if office_lockfile_exists(xlsx_path):
            print(f"\n[XLSX][WARN] {xlsx_path} 正被 WPS/Excel 打开（检测到 ~$ 锁文件），"
                  f"请关闭 WPS/Excel 中打开该文件的窗口后告诉我，我会把数据覆盖写入正式文件。"
                  f"本次结果已写入备用文件 -> {alt}")
        else:
            print(f"\n[XLSX][WARN] {xlsx_path} 被占用（可能被其他程序打开），"
                  f"结果已写入备用文件 -> {alt}")


def office_lockfile_exists(xlsx_path):
    """判断目标 xlsx 是否正被 WPS/Excel 打开：Office 打开文档时会在同目录
    生成 `~$<文件名>`（或去掉扩展名的）隐藏锁文件。"""
    d = os.path.dirname(xlsx_path)
    base = os.path.basename(xlsx_path)
    if not d:
        d = os.getcwd()
    return (os.path.exists(os.path.join(d, "~$" + base)) or
            os.path.exists(os.path.join(d, "~$" + os.path.splitext(base)[0])))

# ============================================================
# 主流程
# ============================================================
def run_single(ser, name, uartbin_path):
    print(f"\n{'='*60}")
    print(f"  Config: {name}")
    print(f"  File:   {uartbin_path}")
    print(f"{'='*60}")

    # 0. 解析体积（ITCM/DTCM 词数，来自 uartbin 帧头）
    itcm_words, dtcm_words = parse_uartbin_size(uartbin_path)

    # 1. 复位
    print("  [1/4] Resetting FPGA...")
    if not reset_fpga():
        return None
    time.sleep(BOOT_DELAY)

    # 2. 下载
    print("  [2/4] Downloading program...")
    ser.reset_input_buffer()
    send_program(ser, uartbin_path)

    # 3. 等待结果
    print("  [3/4] Running CoreMark...")
    text, ok = wait_for_result(ser)

    # 4. 解析
    print("  [4/4] Parsing results...")
    if ok:
        result = parse_coremark(text)
        result['itcm_words'] = itcm_words
        result['dtcm_words'] = dtcm_words
        print(f"    CoreMark/MHz:   {result.get('coremark_mhz', 'N/A')}")
        print(f"    Iterations/Sec: {result.get('iter_per_sec', 'N/A')}")
        print(f"    Total ticks:    {result.get('total_ticks', 'N/A')}")
        print(f"    Branch hit:     {result.get('branch_hit', 'N/A')}")
        return result
    else:
        print(f"    [FAILED] No valid result. Last 200 chars:")
        print(f"    {text[-200:]}")
        return None

def main():
    parser = argparse.ArgumentParser(description="BD32 CoreMark Auto Test")
    parser.add_argument("--port", default=None, help="UART COM port (default: auto-detect CH340)")
    parser.add_argument("--baud", type=int, default=115200, help="Baud rate (default: 115200)")
    parser.add_argument("--compiler", choices=["gcc", "clang"], default=None,
                        help="只跑 GCC 或 LLVM/Clang 编译的配置（默认全部）")
    parser.add_argument("--libc", choices=["newlib", "picolibc"], default=None,
                        help="只跑 newlib-nano 或 picolibc 配置（默认全部）")
    parser.add_argument("--opt", choices=["oz", "os", "o1", "o2", "o3"], default=None,
                        help="只跑指定优化等级（默认全部）")
    parser.add_argument("--lld", action="store_true",
                        help="picolibc LLVM 5 档改用 --lld（LTO+ICF）变体，仅跑这 5 档")
    parser.add_argument("--data-dir", default=None, help="Path to uartbin files")
    parser.add_argument("--xlsx", default=None,
                        help="Excel 输出路径（默认 Working/BD32_CoreMark_Compiler_Comparison.xlsx）")
    args = parser.parse_args()

    # 自动检测串口
    if args.port is None:
        args.port = find_port()
        if args.port is None:
            print("[ERROR] No CH340 USB-UART detected. Use --port to specify manually.")
            sys.exit(1)
        print(f"[AUTO] Detected UART: {args.port}")

    # 确定数据目录
    if args.data_dir:
        data_dir = args.data_dir
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        data_dir = os.path.join(script_dir, "..", "..", "test_data", "soc", "c")
        data_dir = os.path.normpath(data_dir)

    print(f"UART: {args.port} @ {args.baud}")
    print(f"Data: {data_dir}")

    # --lld：只看 picolibc LLVM 5 档的 LTO+ICF 变体（coremark_picolibc_clang_lld_<opt>）
    if args.lld:
        configs = [(name + " (lld)", f"coremark_picolibc_clang_lld_{name.split()[-1].lower()}.uartbin")
                   for name, fname in CONFIGS
                   if name.startswith("Pico LLVM") and "(lld)" not in name]
    else:
        configs = CONFIGS

    # 筛选配置：--compiler / --libc / --opt
    if args.compiler or args.libc or args.opt:
        def match(f):
            if args.compiler:
                is_clang = "clang" in f
                if args.compiler == "gcc" and is_clang:
                    return False
                if args.compiler == "clang" and not is_clang:
                    return False
            if args.libc:
                is_pico = "picolibc" in f
                if args.libc == "newlib" and is_pico:
                    return False
                if args.libc == "picolibc" and not is_pico:
                    return False
            if args.opt and (args.opt + ".uartbin") not in f:
                return False
            return True
        configs = [(n, f) for n, f in CONFIGS if match(f)]
        if not configs:
            print(f"[ERROR] No config matching compiler='{args.compiler}' opt='{args.opt}'")
            sys.exit(1)

    # Excel 输出路径
    if args.xlsx is None:
        args.xlsx = os.path.normpath(os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "..", "..",
            "BD32_CoreMark_Compiler_Comparison.xlsx"))
    else:
        args.xlsx = os.path.abspath(os.path.normpath(args.xlsx))
    # --xlsx 指向已存在目录时，按输出目录处理，拼默认文件名
    if os.path.isdir(args.xlsx):
        args.xlsx = os.path.join(args.xlsx, "BD32_CoreMark_Compiler_Comparison.xlsx")

    # 打开串口
    ser = serial.Serial(args.port, args.baud, timeout=1)
    time.sleep(0.5)

    results = {}
    for name, filename in configs:
        filepath = os.path.join(data_dir, filename)
        if not os.path.exists(filepath):
            print(f"\n[SKIP] {name}: file not found ({filepath})")
            continue
        result = run_single(ser, name, filepath)
        results[name] = result

    ser.close()

    # 汇总
    print(f"\n\n{'='*60}")
    print("  SUMMARY")
    print(f"{'='*60}")
    print(f"{'Config':<12} {'CoreMark/MHz':<14} {'Iter/Sec':<12} {'Ticks':<12} {'Branch%':<10}")
    print("-" * 60)
    for name, _ in configs:
        r = results.get(name)
        if r:
            print(f"{name:<12} {r.get('coremark_mhz',''):<14} {r.get('iter_per_sec',''):<12} "
                  f"{r.get('total_ticks',''):<12} {r.get('branch_hit',''):<10}")
        else:
            print(f"{name:<12} {'FAILED':<14}")

    if results:
        write_xlsx(configs, results, args.xlsx)

if __name__ == "__main__":
    main()
